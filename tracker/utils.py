import textwrap
import openpyxl
import csv
import os
from PIL import Image, ImageDraw, ImageFont, ExifTags, ImageOps
from io import BytesIO
from django.core.files.base import ContentFile
from django.contrib.staticfiles import finders
from geopy.geocoders import Nominatim
import datetime
import functools
from django.core.cache import cache
from django.http import HttpResponseForbidden

# ==========================================
# 5. SECURITY UTILS (NEW)
# ==========================================
def rate_limit(limit=10, period=60):
    """
    Simple IP-based rate limiter decorator.
    limit: Max requests allowed.
    period: Time window in seconds.
    """
    def decorator(view_func):
        @functools.wraps(view_func)
        def _wrapped_view(request, *args, **kwargs):
            ip = request.META.get('REMOTE_ADDR', 'unknown')
            # Create a unique cache key based on view name and IP
            key = f"ratelimit:{view_func.__name__}:{ip}"
            
            # Get current count
            count = cache.get(key, 0)
            
            if count >= limit:
                return HttpResponseForbidden(f"Too many requests. Please try again in {period} seconds.")
            
            # Increment and set/update cache
            if count == 0:
                cache.set(key, 1, period)
            else:
                cache.incr(key)
                
            return view_func(request, *args, **kwargs)
        return _wrapped_view
    return decorator


# ==========================================
# 1. ADDRESS LOOKUP
# ==========================================
def get_address_from_coords(lat, lon):
    if not lat or not lon:
        return "Address Unavailable"
    try:
        # User-agent required by Nominatim
        geolocator = Nominatim(user_agent="tracker_app_v2")
        location = geolocator.reverse((float(lat), float(lon)), exactly_one=True, timeout=5)
        if location:
            return location.address
    except Exception as e:
        print(f"DEBUG: Geocoding Failed: {e}")
    return "Location Unknown"

# ==========================================
# 2. GPS EXTRACTION
# ==========================================
def _convert_to_degrees(value):
    d = float(value[0])
    m = float(value[1])
    s = float(value[2])
    return d + (m / 60.0) + (s / 3600.0)

def get_gps_from_image(image_field):
    try:
        img = Image.open(image_field)
        exif_data = img._getexif()
        if not exif_data: return None, None
        
        gps_info = exif_data.get(34853)
        if not gps_info: return None, None

        lat_gps = gps_info.get(2)
        lat_ref = gps_info.get(1)
        lon_gps = gps_info.get(4)
        lon_ref = gps_info.get(3)

        if lat_gps and lat_ref and lon_gps and lon_ref:
            lat = _convert_to_degrees(lat_gps)
            if lat_ref != 'N': lat = -lat
            lon = _convert_to_degrees(lon_gps)
            if lon_ref != 'E': lon = -lon
            return f"{lat:.6f}", f"{lon:.6f}"
    except Exception:
        pass
    return None, None

# ==========================================
# 3. DYNAMIC WATERMARKING LOGIC
# ==========================================
def watermark_image(image_field, lat, lon):
    try:
        print("DEBUG: Processing Image for Watermark...")
        
        COMPANY_NAME = "Nexsafe"
        
        # 1. Prepare Data
        if lat and lon:
            address_text = get_address_from_coords(lat, lon)
            gps_text = f"Lat: {float(lat):.6f}, Lon: {float(lon):.6f}"
        else:
            address_text = "Location Not Captured"
            gps_text = "GPS Unavailable"

        # 2. Load & Orient Image
        img = Image.open(image_field)
        img = ImageOps.exif_transpose(img) # Critical for phone photos!
        img = img.convert("RGBA")
        draw = ImageDraw.Draw(img)
        W, H = img.size
        
        # --- DYNAMIC SIZING CALCULATIONS ---
        # We base everything on the SHORTEST side to ensure consistency 
        # whether the photo is Portrait or Landscape.
        base_dim = min(W, H)
        
        # Configurable Ratios
        LOGO_RATIO = 0.2    # Logo is 12% of shortest side (Smaller than before)
        TITLE_RATIO = 0.05   # Title is 5% (Larger than before)
        BODY_RATIO = 0.035   # Body is 3.5% (Readable on phones)
        PADDING_RATIO = 0.02 # Padding is 2%

        logo_target_size = int(base_dim * LOGO_RATIO)
        font_title_size = int(base_dim * TITLE_RATIO)
        font_body_size = int(base_dim * BODY_RATIO)
        padding = int(base_dim * PADDING_RATIO)

        # 3. Load Fonts
        # Try multiple common font paths for Linux/Windows servers
        font_paths = [
            "arial.ttf", 
            "DejaVuSans-Bold.ttf", 
            "DejaVuSans.ttf",
            "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"
        ]
        
        font_title = None
        font_body = None

        for path in font_paths:
            try:
                if not font_title: font_title = ImageFont.truetype(path, font_title_size)
                if not font_body: font_body = ImageFont.truetype(path, font_body_size)
            except OSError: continue
            
        # Fallback if no TTF found (Default font is tiny, but better than crash)
        if not font_title: font_title = ImageFont.load_default()
        if not font_body: font_body = ImageFont.load_default()

        # 4. Wrap Address (approx characters based on width)
        # We calculate wrap width dynamically based on font size
        chars_per_line = 40 
        wrapped_address = "\n".join(textwrap.wrap(address_text, width=chars_per_line))

        # 5. Load Logo
        logo_path = finders.find('tracker/logo.png')
        logo = None
        if logo_path:
            try:
                logo = Image.open(logo_path).convert("RGBA")
                # Resize logo maintaining aspect ratio
                aspect = logo.width / logo.height
                new_h = logo_target_size
                new_w = int(new_h * aspect)
                logo.thumbnail((new_w, new_h), Image.Resampling.LANCZOS)
            except Exception: pass

        # 6. Measure Text Block
        def get_text_size(text, font):
            if hasattr(draw, "textbbox"):
                bbox = draw.textbbox((0, 0), text, font=font)
                return bbox[2], bbox[3]
            else:
                return draw.textsize(text, font=font)

        w_t, h_t = get_text_size(COMPANY_NAME, font_title)
        w_g, h_g = get_text_size(gps_text, font_body)
        w_a, h_a = get_text_size(wrapped_address, font_body)

        text_width = max(w_t, w_g, w_a)
        # Add line spacings
        text_height = h_t + h_g + h_a + (padding * 1.5)

        # 7. Calculate Box Dimensions
        logo_w = logo.size[0] if logo else 0
        logo_h = logo.size[1] if logo else 0
        
        box_width = logo_w + text_width + (padding * 3)
        # Ensure box is tall enough for either text OR logo
        box_height = max(text_height, logo_h) + (padding * 2)

        # 8. Placement (Bottom Right)
        x2 = W - padding
        y2 = H - padding
        x1 = x2 - box_width
        y1 = y2 - box_height

        # 9. Draw Background Box
        overlay = Image.new("RGBA", img.size, (0, 0, 0, 0))
        overlay_draw = ImageDraw.Draw(overlay)
        
        # Rounded corners if supported, else rectangle
        if hasattr(overlay_draw, "rounded_rectangle"):
            overlay_draw.rounded_rectangle([x1, y1, x2, y2], radius=int(padding/2), fill=(0, 0, 0, 180))
        else:
            overlay_draw.rectangle([x1, y1, x2, y2], fill=(0, 0, 0, 180))
        
        img = Image.alpha_composite(img, overlay)
        draw = ImageDraw.Draw(img)

        # 10. Draw Content
        current_x = x1 + padding
        current_y = y1 + padding

        # Draw Logo (Centered Vertically in Box)
        if logo:
            logo_y = y1 + (box_height - logo_h) // 2
            img.paste(logo, (int(current_x), int(logo_y)), logo)
            current_x += logo_w + padding

        # Draw Text
        # Title
        draw.text((current_x, current_y), COMPANY_NAME, fill="white", font=font_title)
        current_y += h_t + (padding * 0.2)
        
        # GPS
        draw.text((current_x, current_y), gps_text, fill="#d0d0d0", font=font_body)
        current_y += h_g + (padding * 0.2)
        
        # Address
        draw.text((current_x, current_y), wrapped_address, fill="#b0b0b0", font=font_body)

        # 11. Output
        buffer = BytesIO()
        img.convert("RGB").save(buffer, format='JPEG', quality=95)
        return ContentFile(buffer.getvalue())

    except Exception as e:
        print(f"DEBUG: Watermark Logic Crashed: {e}")
        image_field.seek(0)
        return image_field

# ==========================================
# 4. EXCEL HELPERS
# ==========================================
def get_file_headers(file_field):
    if not file_field: return []
    try:
        try: file_field.open('rb')
        except: pass
        file_field.seek(0)
        filename = file_field.name.lower()
        if filename.endswith('.xlsx'):
            workbook = openpyxl.load_workbook(file_field, data_only=True)
            sheet = workbook.active
            return [str(cell.value).strip() for cell in sheet[1] if cell.value]
        else:
            decoded = file_field.read().decode('utf-8-sig').splitlines()
            reader = csv.reader(decoded)
            return next(reader, [])
    except: return []

def get_dropdown_options(file_field, column_name):
    if not file_field: return []
    options = set()
    try:
        try: file_field.open('rb')
        except: pass
        file_field.seek(0)
        filename = file_field.name.lower()
        if filename.endswith('.xlsx'):
            workbook = openpyxl.load_workbook(file_field, data_only=True)
            sheet = workbook.active
            headers = [str(cell.value).strip() if cell.value else '' for cell in sheet[1]]
            try:
                idx = headers.index(column_name.strip())
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if row[idx]: options.add(str(row[idx]).strip())
            except ValueError: pass
        else:
            decoded = file_field.read().decode('utf-8-sig').splitlines()
            reader = csv.DictReader(decoded)
            reader.fieldnames = [name.strip() for name in reader.fieldnames]
            if column_name.strip() in reader.fieldnames:
                for row in reader:
                    if row.get(column_name.strip()): options.add(row.get(column_name.strip()).strip())
        return [(o, o) for o in sorted(list(options))]
    except: return []